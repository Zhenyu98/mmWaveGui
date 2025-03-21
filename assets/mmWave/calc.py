import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import animation
from openpyxl import Workbook

with open(r"E:\Perception Neuron\Record\转身_Char00.calc", 'r') as file:
    for _ in range(6):
        next(file)

    file_content = file.read()
    

file_content=file_content.split()
data_num=len(file_content)
frames_num=data_num/946
frames_num=int(frames_num)
print(frames_num)

data=[[] for _ in range(frames_num)]
m=0
for i in range(frames_num):
    for j in range(336): #21*16

        data[i].append(float(file_content[m*946+j]))
    m=m+1



fig = plt.figure()
ax=fig.add_subplot(111,projection='3d')
line, = ax.plot([], [], [], 'r-')

ax.set_xlim(-1,1)
ax.set_ylim(-1,1)
ax.set_zlim(-1.5,1.5)

skeleton=[
    [], #[[x0,y0,z0],[x1,y1,z1],...]
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [] #21
]


for i in range(21):

    skeleton[i]=[[] for _ in range(frames_num)]

k=0


for i in range(21):
    for j in range(frames_num):

        skeleton[i][j].append(data[j][k])
        skeleton[i][j].append(data[j][k+1])
        skeleton[i][j].append(data[j][k+2])

    k=k+16

#skeleton=[[[-x for x in frame] for frame in sublink] for sublink in skeleton]


#数据保存


workbook = Workbook()


sheet = workbook.active


sheet['A1'] = 'Hips_X'
sheet['B1'] = 'Hips_Y'
sheet['C1'] = 'Hips_Z'
sheet['D1'] = 'RightUpLeg_X'
sheet['E1'] = 'RightUpLeg_Y'
sheet['F1'] = 'RightUpLeg_Z'
sheet['G1'] = 'RightLeg_X'
sheet['H1'] = 'RightLeg_Y'
sheet['I1'] = 'RightLeg_Z'
sheet['J1'] = 'RightFoot_X'
sheet['K1'] = 'RightFoot_Y'
sheet['L1'] = 'RightFoot_Z'
sheet['M1'] = 'LeftUpLeg_X'
sheet['N1'] = 'LeftUpLeg_Y'
sheet['O1'] = 'LeftUpLeg_Z'
sheet['P1'] = 'LeftLeg_X'
sheet['Q1'] = 'LeftLeg_Y'
sheet['R1'] = 'LeftLeg_Z'
sheet['S1'] = 'LeftFoot_X'
sheet['T1'] = 'LeftFoot_Y'
sheet['U1'] = 'LeftFoot_Z'
sheet['V1'] = 'RightShoulder_X'
sheet['W1'] = 'RightShoulder_Y'
sheet['X1'] = 'RightShoulder_Z'
sheet['Y1'] = 'RightArm_X'
sheet['Z1'] = 'RightArm_Y'
sheet['AA1'] = 'RightArm_Z'
sheet['AB1'] = 'RightForeArm_X'
sheet['AC1'] = 'RightForeArm_Y'
sheet['AD1'] = 'RightForeArm_Z'
sheet['AE1'] = 'RightHand_X'
sheet['AF1'] = 'RightHand_Y'
sheet['AG1'] = 'RightHand_Z'
sheet['AH1'] = 'LeftShoulder_X'
sheet['AI1'] = 'LeftShoulder_Y'
sheet['AJ1'] = 'LeftShoulder_Z'
sheet['AK1'] = 'LeftArm_X'
sheet['AL1'] = 'LeftArm_Y'
sheet['AM1'] = 'LeftArm_Z'
sheet['AN1'] = 'LeftForeArm_X'
sheet['AO1'] = 'LeftForeArm_Y'
sheet['AP1'] = 'LeftForeArm_Z'
sheet['AQ1'] = 'LeftHand_X'
sheet['AR1'] = 'LeftHand_Y'
sheet['AS1'] = 'LeftHand_Z'
sheet['AT1'] = 'Head_X'
sheet['AU1'] = 'Head_Y'
sheet['AV1'] = 'Head_Z'
sheet['AW1'] = 'Neck_X'
sheet['AX1'] = 'Neck_Y'
sheet['AY1'] = 'Neck_Z'
sheet['AZ1'] = 'Spine3_X'
sheet['BA1'] = 'Spine3_Y'
sheet['BB1'] = 'Spine3_Z'
sheet['BC1'] = 'Spine2_X'
sheet['BD1'] = 'Spine2_Y'
sheet['BE1'] = 'Spine2_Z'
sheet['BF1'] = 'Spine1_X'
sheet['BG1'] = 'Spine1_Y'
sheet['BH1'] = 'Spine1_Z'
sheet['BI1'] = 'Spine_X'
sheet['BJ1'] = 'Spine_Y'
sheet['BK1'] = 'Spine_Z'

for i in range(frames_num):
    k=0
    for j in range(21):
        sheet.cell(row=i + 2, column=k + 1, value=skeleton[j][i][0])
        sheet.cell(row=i + 2, column=k + 2, value=skeleton[j][i][1])
        sheet.cell(row=i + 2, column=k + 3, value=skeleton[j][i][2])
        k=k+3


workbook.save('E:/Perception Neuron/data/转身_calc_data_processing.xlsx')



head=[0,0,0]
neck=[0,0,0]
leftshouder=[0,0,0]
leftarm=[0,0,0]
leftforearm=[0,0,0]
lefthand=[0,0,0]
rightshouder=[0,0,0]
rightarm=[0,0,0]
rightforearm=[0,0,0]
righthand=[0,0,0]
spine3=[0,0,0]
spine2=[0,0,0]
spine1=[0,0,0]
spine=[0,0,0]
hips=[0,0,0]
rightupleg=[0,0,0]
rightleg=[0,0,0]
rightfoot=[0,0,0]
leftupleg=[0,0,0]
leftleg=[0,0,0]
leftfoot=[0,0,0]

def update(frames_num):
    for artist in ax.lines + ax.collections:
        artist.remove()


    for i in range(3):
        head[i] = skeleton[15][frames_num][i]
        neck[i] = skeleton[16][frames_num][i]
        leftshouder[i] = skeleton[11][frames_num][i]
        leftarm[i] = skeleton[12][frames_num][i]
        leftforearm[i] = skeleton[13][frames_num][i]
        lefthand[i] = skeleton[14][frames_num][i]
        rightshouder[i] = skeleton[7][frames_num][i]
        rightarm[i] = skeleton[8][frames_num][i]
        rightforearm[i] = skeleton[9][frames_num][i]
        righthand[i] = skeleton[10][frames_num][i]
        spine3[i] = skeleton[17][frames_num][i]
        spine2[i] = skeleton[18][frames_num][i]
        spine1[i] = skeleton[19][frames_num][i]
        spine[i] = skeleton[20][frames_num][i]
        hips[i] = skeleton[0][frames_num][i]
        rightupleg[i] = skeleton[1][frames_num][i]
        rightleg[i] = skeleton[2][frames_num][i]
        rightfoot[i] = skeleton[3][frames_num][i]
        leftupleg[i] = skeleton[4][frames_num][i]
        leftleg[i] = skeleton[5][frames_num][i]
        leftfoot[i]=skeleton[6][frames_num][i]

    ax.plot([head[0],neck[0]],
            [head[1],neck[1]],
            [head[2],neck[2]],'green')

    ax.plot([neck[0],leftshouder[0],leftarm[0],leftforearm[0],lefthand[0]],
            [neck[1],leftshouder[1],leftarm[1],leftforearm[1],lefthand[1]],
            [neck[2],leftshouder[2],leftarm[2],leftforearm[2],lefthand[2]],'red')

    ax.plot([neck[0], rightshouder[0], rightarm[0], rightforearm[0], righthand[0]],
            [neck[1], rightshouder[1], rightarm[1], rightforearm[1], righthand[1]],
            [neck[2], rightshouder[2], rightarm[2], rightforearm[2], righthand[2]], 'blue')

    ax.plot([neck[0],spine3[0],spine2[0],spine1[0],spine[0],hips[0]],
            [neck[1],spine3[1],spine2[1],spine1[1],spine[1],hips[1]],
            [neck[2],spine3[2],spine2[2],spine1[2],spine[2],hips[2]],'black')

    ax.plot([hips[0], rightupleg[0], rightleg[0], rightfoot[0]],
            [hips[1], rightupleg[1], rightleg[1], rightfoot[1]],
            [hips[2], rightupleg[2], rightleg[2], rightfoot[2]], 'purple')

    ax.plot([hips[0], leftupleg[0], leftleg[0], leftfoot[0]],
            [hips[1], leftupleg[1], leftleg[1], leftfoot[1]],
            [hips[2], leftupleg[2], leftleg[2], leftfoot[2]], 'orange')





    return ax


def init():

    line.set_data([], [])
    line.set_3d_properties([])
    return line,


ani = animation.FuncAnimation(fig,update,int(frames_num),init_func = init,interval=20)
plt.show()





