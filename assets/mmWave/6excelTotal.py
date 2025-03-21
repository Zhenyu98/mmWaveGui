import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# 文件夹路径
base_folder_path = r'C:\Users\Kano\Desktop\radar_toolbox_1_30_01_03\tools\visualizers\Industrial_Visualizer\binData\traindata\9sign'

# 文件夹数量
folder_count = 50

# 创建新的Excel工作簿
combined_workbook = Workbook()
combined_sheet = combined_workbook.active

# 遍历文件夹
for i in range(1, folder_count + 1):
    folder_path = os.path.join(base_folder_path, str(i))
    print(f'读取Excel文件 {i}')

    # 创建一个空的列表来存储每个子文件夹的数据
    data_lists = []

    # 遍历文件夹后缀
    for suffix in ['pHistBytes_clustered_voxel_XOZ', 'pHistBytes_clustered_voxel_YOZ']:
        # 文件夹完整路径
        full_folder_path = os.path.join(folder_path, suffix)

        # 检查子文件夹中的Excel文件是否存在
        excel_file_path = os.path.join(full_folder_path, 'image_data.xlsx')
        if not os.path.exists(excel_file_path):
            continue

        # 打开Excel文件
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # 读取数据
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # 将数据添加到列表中
        data_lists.append(data)

    # 检查两个子文件夹的数据是否都存在
    if len(data_lists) == 2:
        # 写入表头
        if i == 1:
            combined_sheet.append(data_lists[0][0] + data_lists[1][0])

        # 拼接数据到新的Excel工作簿中（不包括表头）
        for j in range(1, len(data_lists[0])):
            combined_sheet.append(data_lists[0][j] + data_lists[1][j])

# 更新列宽
for column in combined_sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.2
    combined_sheet.column_dimensions[column_letter].width = adjusted_width

# 保存合并后的Excel文件
combined_file_path = os.path.join(base_folder_path, 'combined_image_data.xlsx')
combined_workbook.save(combined_file_path)

print('合并后的Excel文件已保存：', combined_file_path)